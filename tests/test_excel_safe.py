"""Union regression suite — every grafted fix from the 2026-06-11 diff matrix."""

import io
import zipfile

import openpyxl
import pytest

from tropi_common.excel_safe import SafeWorkbook


def _make_xlsx(with_formula=True) -> bytes:
    """Build a fixture and normalize it to Excel's canonical serialization.

    openpyxl writes `<sheet xmlns:r="..." name=...>` (inline namespace first),
    which Excel never produces; the production sheet parser — byte-identical
    across all 8 service copies and deliberately untouched by the union — only
    needs to handle Excel-authored files. Fixtures must mimic those.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Кg "          # trailing space — the wr regression value
    ws["B1"] = 42
    if with_formula:
        ws["C1"] = "=B1*2"
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    # rewrite workbook.xml: hoist xmlns:r to the root (Excel-canonical)
    src = io.BytesIO(raw)
    out = io.BytesIO()
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/workbook.xml":
                xml = data.decode()
                xml = xml.replace(' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"', "", xml.count("<sheet "))
                if "xmlns:r=" not in xml.split(">", 1)[0] + ">":
                    xml = xml.replace("<workbook ", '<workbook xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ', 1)
                data = xml.encode()
            zout.writestr(item, data)
    return out.getvalue()


# ── wr graft: cell_str whitespace preserve ───────────────────────────────────

def test_cell_str_preserves_trailing_whitespace():
    xml = SafeWorkbook.cell_str("A", 1, "Кg ")
    assert 'xml:space="preserve"' in xml
    assert ">Кг ".replace("г", "g") in xml or ">Кg <" in xml  # value intact incl. space


def test_cell_str_no_preserve_attr_for_clean_strings():
    xml = SafeWorkbook.cell_str("A", 1, "Кg")
    assert "xml:space" not in xml


def test_trailing_space_survives_save_roundtrip():
    sw = SafeWorkbook(_make_xlsx())
    sheet = sw.sheet_xml("Sheet1")
    sheet = sw.inject_cell(sheet, "D", 1, SafeWorkbook.cell_str("D", 1, "Кg "))
    sw.set_sheet_xml("Sheet1", sheet)
    out = sw.save()
    wb = openpyxl.load_workbook(io.BytesIO(out))
    assert wb["Sheet1"]["D1"].value == "Кg "  # 4 chars, space intact


# ── keyaccounts graft: 3-tier inject_cell ────────────────────────────────────

def test_inject_cell_creates_missing_row():
    sw = SafeWorkbook(_make_xlsx())
    sheet = sw.sheet_xml("Sheet1")
    assert '<row r="34"' not in sheet
    sheet = sw.inject_cell(sheet, "B", 34, SafeWorkbook.cell_str("B", 34, "подпис"))
    assert '<row r="34">' in sheet
    sw.set_sheet_xml("Sheet1", sheet)
    wb = openpyxl.load_workbook(io.BytesIO(sw.save()))
    assert wb["Sheet1"]["B34"].value == "подпис"


# ── keyaccounts graft: EMF guard in save ─────────────────────────────────────

def test_save_keeps_emf_default_while_emf_parts_remain():
    src = _make_xlsx()
    sw = SafeWorkbook(src)
    # plant an emf part + its Content-Types Default
    sw._data["xl/media/image1.emf"] = b"emfbytes"
    import zipfile as zf_mod
    sw._infos["xl/media/image1.emf"] = zf_mod.ZipInfo("xl/media/image1.emf")
    ct = sw._data["[Content_Types].xml"].decode()
    ct = ct.replace("<Types ", "<Types ", 1).replace(
        "</Types>", '<Default Extension="emf" ContentType="image/x-emf"/></Types>')
    sw._data["[Content_Types].xml"] = ct.encode()
    out = sw.save()
    with zipfile.ZipFile(io.BytesIO(out)) as z:
        assert 'Extension="emf"' in z.read("[Content_Types].xml").decode()


def test_save_strips_emf_default_when_no_emf_parts():
    src = _make_xlsx()
    sw = SafeWorkbook(src)
    ct = sw._data["[Content_Types].xml"].decode().replace(
        "</Types>", '<Default Extension="emf" ContentType="image/x-emf"/></Types>')
    sw._data["[Content_Types].xml"] = ct.encode()
    out = sw.save()
    with zipfile.ZipFile(io.BytesIO(out)) as z:
        assert 'Extension="emf"' not in z.read("[Content_Types].xml").decode()


# ── wr graft: shift_formula_refs exists and shifts array refs ────────────────

def test_shift_formula_refs_moves_array_anchor():
    sheet = '<row r="5"><c r="A5"><f t="array" ref="A5:A7">SUM(B5:B7)</f></c></row>'
    out = SafeWorkbook.shift_formula_refs(sheet, 5, 3)
    assert 'ref="A8:A10"' in out


# ── keyaccounts variant: shared-formula expansion incl. self-closing <v/> ────

def test_expand_shared_formulas_handles_self_closing_v():
    sheet = (
        '<sheetData>'
        '<row r="1"><c r="A1"><f ref="A1:A2" t="shared" si="0">B1*2</f><v>4</v></c></row>'
        '<row r="2"><c r="A2"><f t="shared" si="0"/><v/></c></row>'
        '</sheetData>'
    )
    out = SafeWorkbook.expand_shared_formulas(sheet)
    assert 't="shared"' not in out
    assert "<f>B2*2</f>" in out


# ── golden-file round-trip ───────────────────────────────────────────────────

def test_golden_roundtrip_only_documented_deltas():
    """Open → touch nothing → save. Every ZIP member must be byte-identical
    except save()'s documented transforms: calcChain dropped, workbook.xml
    gains fullCalcOnLoad, [Content_Types].xml may lose dropped parts."""
    src = _make_xlsx(with_formula=True)
    out = SafeWorkbook(src).save()
    with zipfile.ZipFile(io.BytesIO(src)) as a, zipfile.ZipFile(io.BytesIO(out)) as b:
        names_a, names_b = set(a.namelist()), set(b.namelist())
        assert names_a - names_b <= {"xl/calcChain.xml"}
        assert names_b <= names_a
        ALLOWED_DIFF = {"xl/workbook.xml", "[Content_Types].xml"}
        for name in names_b:
            if a.read(name) != b.read(name):
                assert name in ALLOWED_DIFF, f"unexpected mutation in {name}"
        wb = openpyxl.load_workbook(io.BytesIO(out))
        assert wb["Sheet1"]["B1"].value == 42
